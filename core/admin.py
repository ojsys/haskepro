from django.contrib import admin
from django.http import HttpResponse
from django.urls import path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from .models import HeroSlide, Statistics, Achievement, Ministry, MinistrySection, Gallery, DemographicData, SiteLogo, Subscriber, AboutPage, AboutMinistry, MissionVision, Challenge, BoardMember
from .models import Project, ProjectPage, VolunteerPage, GoTeam, GiveSection, PrayerPartner, VolunteerApplication
from .models import BlogPost, YouTubeVideo, SpotifyPodcast, MediaPage, DonationPage, BankAccount
from import_export.admin import ImportExportModelAdmin
from import_export.formats import base_formats
import xlsxwriter
from .resources import DemographicDataResource


admin.site.register(SiteLogo)

@admin.register(HeroSlide)
class HeroSlideAdmin(admin.ModelAdmin):
    list_display = ('title', 'order', 'is_active', 'created_at')
    list_editable = ('order', 'is_active')
    search_fields = ('title', 'description')
    list_filter = ('is_active',)
    ordering = ('order',)

@admin.register(Statistics)
class StatisticsAdmin(admin.ModelAdmin):
    list_display = ('title', 'total_population', 'people_groups', 'villages_reached', 'converts', 'film_attendees', 'people_reached', 'updated_at')

@admin.register(Achievement)
class AchievementAdmin(admin.ModelAdmin):
    list_display = ('villages', 'people', 'reached', 'updated_at')

    def has_add_permission(self, request):
        # Only allow one instance
        if self.model.objects.exists():
            return False
        return True


@admin.register(Ministry)
class MinistryAdmin(admin.ModelAdmin):
    list_display = ('title', 'order')
    list_editable = ('order',)
    search_fields = ('title', 'description')
    ordering = ('order',)

@admin.register(Gallery)
class GalleryAdmin(admin.ModelAdmin):
    list_display = ('title', 'is_featured', 'created_at')
    list_editable = ('is_featured',)
    search_fields = ('title', 'description')
    list_filter = ('is_featured',)
    

@admin.register(MinistrySection)
class MinistrySectionAdmin(admin.ModelAdmin):
    list_display = ('title', 'description', 'image')

    def has_add_permission(self, request):
        # Only allow one instance
        if self.model.objects.exists():
            return False
        return True


@admin.register(DemographicData)
class DemographicDataAdmin(ImportExportModelAdmin):
    resources_class = DemographicDataResource
    formats = [base_formats.XLSX]

    list_display = ['state', 'lga', 'ward', 'village', 'total_village_population', 'converts']
    list_filter = ['state', 'lga']
    search_fields = ['state', 'lga', 'ward', 'village']
    actions = ['export_selected_state']
    
    fieldsets = (
        ('Location', {
            'fields': ('state', 'lga', 'ward', 'village')
        }),
        ('Population Data', {
            'fields': (
                'christian_population',
                'muslim_population',
                'traditional_population',
                'converts',
                'total_village_population'
            )
        }),
        ('Additional Information', {
            'fields': ('film_attendance', 'people_group', 'practiced_religion')
        })
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('download-template/', 
                 self.download_template, 
                 name='demographic-data-template'),
        ]
        return custom_urls + urls

    def download_template(self, request):
        # Create the HttpResponse object with Excel header
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="demographic_data_template.xlsx"'

        # Create the Excel workbook and add a worksheet
        workbook = xlsxwriter.Workbook(response)
        worksheet = workbook.add_worksheet()

        # Add headers
        headers = ['state', 'village', 'total_village_population', 'converts']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        # Add some example data (optional)
        example_data = [
            ['Lagos', 'Village A', 1000, 50],
            ['Abuja', 'Village B', 2000, 100],
        ]
        for row, data in enumerate(example_data, start=1):
            for col, value in enumerate(data):
                worksheet.write(row, col, value)

        workbook.close()
        return response

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_template_download'] = True
        return super().changelist_view(request, extra_context)

    # def get_urls(self):
    #     urls = super().get_urls()
    #     custom_urls = [
    #         path('export-state/<str:state_name>/', 
    #              self.admin_site.admin_view(self.export_state_data), 
    #              name='export-state-data'),
    #     ]
    #     return custom_urls + urls

    def export_selected_state(self, request, queryset):
        # Get unique state from selection
        state_name = queryset.first().state if queryset.exists() else "Unknown"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{state_name} State Data"
        
        # Define headers
        headers = [
            'LGA', 'Ward', 'Village', 'Population', 'Christians',
            'Muslims', 'Traditional', 'Converts', 'Film Attendance'
        ]
        
        # Style headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        for row, data in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=data.lga)
            ws.cell(row=row, column=2, value=data.ward)
            ws.cell(row=row, column=3, value=data.village)
            ws.cell(row=row, column=4, value=data.total_village_population)
            ws.cell(row=row, column=5, value=data.christian_population)
            ws.cell(row=row, column=6, value=data.muslim_population)
            ws.cell(row=row, column=7, value=data.traditional_population)
            ws.cell(row=row, column=8, value=data.converts)
            ws.cell(row=row, column=9, value=data.film_attendance)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{state_name}_data_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        wb.save(response)
        return response
    
    export_selected_state.short_description = "Export selected data to Excel"

    def import_action(self, request, *args, **kwargs):
        """
        Custom import action to handle errors better
        """
        try:
            return super().import_action(request, *args, **kwargs)
        except Exception as e:
            self.message_user(request, f"Import Error: {str(e)}", level=messages.ERROR)
            return HttpResponseRedirect("../")


@admin.register(Subscriber)
class SubscriberAdmin(admin.ModelAdmin):
    list_display = ('email', 'created_at', 'is_active')
    list_filter = ('is_active', 'created_at')
    search_fields = ('email',)
    date_hierarchy = 'created_at'

###### ABout ######
@admin.register(AboutPage)
class AboutPageAdmin(admin.ModelAdmin):
    def has_add_permission(self, request):
        # Limit to single instance
        if self.model.objects.exists():
            return False
        return True

@admin.register(AboutMinistry)
class AboutMinistryAdmin(admin.ModelAdmin):
    list_display = ('title', 'order')
    list_editable = ('order',)

@admin.register(MissionVision)
class MissionVisionAdmin(admin.ModelAdmin):
    def has_add_permission(self, request):
        # Limit to single instance
        if self.model.objects.exists():
            return False
        return True

@admin.register(Challenge)
class ChallengeAdmin(admin.ModelAdmin):
    list_display = ('title', 'order')
    list_editable = ('order',)
    fieldsets = (
        (None, {
            'fields': ('title', 'description', 'order')
        }),
        ('Challenge 1', {
            'fields': ('challenge1_title', 'challenge1_desc', 'challenge1_image')
        }),
        ('Challenge 2', {
            'fields': ('challenge2_title', 'challenge2_desc', 'challenge2_image')
        }),
        ('Challenge 3', {
            'fields': ('challenge3_title', 'challenge3_desc', 'challenge3_image')
        }),
    )

@admin.register(BoardMember)
class BoardMemberAdmin(admin.ModelAdmin):
    list_display = ('name', 'position', 'order')
    list_editable = ('order',)

#########  PROJECTs ###########

@admin.register(ProjectPage)
class ProjectPageAdmin(admin.ModelAdmin):
    def has_add_permission(self, request):
        if self.model.objects.exists():
            return False
        return True

@admin.register(Project)
class ProjectAdmin(admin.ModelAdmin):
    list_display = ('title', 'order')
    list_editable = ('order',)

##############################

############  Volunteer ################
@admin.register(VolunteerPage)
class VolunteerPageAdmin(admin.ModelAdmin):
    def has_add_permission(self, request):
        if self.model.objects.exists():
            return False
        return True

@admin.register(VolunteerApplication)
class VolunteerApplicationAdmin(admin.ModelAdmin):
    list_display = ('full_name', 'email', 'area_of_interest', 'created_at')
    list_filter = ('area_of_interest', 'created_at')
    search_fields = ('full_name', 'email')
    readonly_fields = ('created_at',)


admin.site.register(GoTeam)
admin.site.register(PrayerPartner)
admin.site.register(GiveSection)

########################################

################# MEDIA ADMIN #########

@admin.register(BlogPost)
class BlogPostAdmin(admin.ModelAdmin):
    list_display = ('title', 'author', 'published_date', 'is_featured')
    list_filter = ('is_featured', 'published_date')
    search_fields = ('title', 'content')
    prepopulated_fields = {'slug': ('title',)}

@admin.register(YouTubeVideo)
class YouTubeVideoAdmin(admin.ModelAdmin):
    list_display = ('title', 'published_date', 'is_featured')
    list_filter = ('is_featured', 'published_date')
    search_fields = ('title', 'description')

@admin.register(SpotifyPodcast)
class SpotifyPodcastAdmin(admin.ModelAdmin):
    list_display = ('title', 'published_date', 'duration', 'is_featured')
    list_filter = ('is_featured', 'published_date')
    search_fields = ('title', 'description')

@admin.register(MediaPage)
class MediaPageAdmin(admin.ModelAdmin):
    def has_add_permission(self, request):
        if self.model.objects.exists():
            return False
        return True
    
#########################################
@admin.register(DonationPage)
class DonationPageAdmin(admin.ModelAdmin):
    list_display = ('title', 'bank_name')
    
    def has_add_permission(self, request):
        # Only allow one instance
        if self.model.objects.exists():
            return False
        return True

@admin.register(BankAccount) 
class BankAccountAdmin(admin.ModelAdmin):
    list_display = ('account_name', 'account_number', 'description')
    list_editable = ('account_number', 'description')
    search_fields = ('account_name', 'account_number')
    ordering = ('account_name',)